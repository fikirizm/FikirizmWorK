import uuid
import os
import re
import io
import logging
import httpx
from datetime import datetime, timezone, timedelta
from fastapi import APIRouter, Request, Response, HTTPException, Depends, WebSocket, WebSocketDisconnect, Query
from fastapi.responses import StreamingResponse
from typing import Optional

from deps import (
    db, get_current_user, get_ws_user, hash_password, verify_password,
    create_access_token, create_refresh_token, set_auth_cookies, clear_auth_cookies,
    sanitize_user,
)
from realtime import manager
from models import (
    RegisterBody, LoginBody, SessionBody, WorkspaceBody, ProjectBody, ProjectUpdate,
    TaskBody, TaskUpdate, BulkUpdate, CommentBody, IdeaBody, IdeaUpdate, InviteBody,
    BudgetBody, BudgetUpdate,
)
from templates_data import TEMPLATES, get_template
from mailer import email_task_assigned, email_project_added

logger = logging.getLogger("fikirizm.routes")

router = APIRouter(prefix="/api")

DEFAULT_STATUSES = [
    {"id": "todo", "name": "Yapılacak", "color": "#71717A", "order": 0},
    {"id": "in_progress", "name": "Devam Ediyor", "color": "#3B82F6", "order": 1},
    {"id": "review", "name": "İncelemede", "color": "#F59E0B", "order": 2},
    {"id": "done", "name": "Tamamlandı", "color": "#10B981", "order": 3, "done": True},
]


def is_privileged(user):
    return user.get("role") in ("owner", "admin")


def can_access_project(user, project):
    if not project:
        return False
    if is_privileged(user):
        return True
    return user["user_id"] in project.get("members", []) or project.get("created_by") == user["user_id"]


def can_edit_budget(user, project):
    if is_privileged(user):
        return True
    if project.get("budget_policy") == "members":
        return can_access_project(user, project)
    return False


async def accessible_project_ids(user):
    if is_privileged(user):
        return None  # None => all projects in org
    projs = await db.projects.find(
        {"org_id": user["org_id"]}, {"_id": 0, "id": 1, "members": 1, "created_by": 1}).to_list(2000)
    return [p["id"] for p in projs if user["user_id"] in p.get("members", []) or p.get("created_by") == user["user_id"]]


def can_see_task(user, task):
    if is_privileged(user):
        return True
    if task.get("visibility") == "private":
        return (user["user_id"] in (task.get("visible_to") or [])
                or user["user_id"] in task.get("assignees", [])
                or task.get("created_by") == user["user_id"])
    return True


async def _email_new_assignees(new_ids, actor, task_title, project_name=""):
    for aid in new_ids:
        if aid == actor["user_id"]:
            continue
        u = await db.users.find_one({"user_id": aid}, {"_id": 0})
        if u and u.get("email"):
            await email_task_assigned(u["email"], u.get("name", ""), actor.get("name", ""), task_title, project_name)


async def _email_new_members(new_ids, actor, project_name):
    for aid in new_ids:
        if aid == actor["user_id"]:
            continue
        u = await db.users.find_one({"user_id": aid}, {"_id": 0})
        if u and u.get("email"):
            await email_project_added(u["email"], u.get("name", ""), actor.get("name", ""), project_name)


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def new_id(prefix=""):
    return f"{prefix}{uuid.uuid4().hex[:16]}"


async def log_activity(org_id, workspace_id, user, action, target=""):
    doc = {
        "id": new_id("act_"),
        "org_id": org_id,
        "workspace_id": workspace_id,
        "user_id": user["user_id"],
        "user_name": user.get("name", ""),
        "action": action,
        "target": target,
        "created_at": now_iso(),
    }
    await db.activities.insert_one(doc)
    await manager.broadcast(workspace_id, {"type": "activity", "data": {k: v for k, v in doc.items() if k != "_id"}})


async def create_notification(org_id, user_id, ntype, message, link=""):
    doc = {
        "id": new_id("ntf_"),
        "org_id": org_id,
        "user_id": user_id,
        "type": ntype,
        "message": message,
        "link": link,
        "read": False,
        "created_at": now_iso(),
    }
    await db.notifications.insert_one(doc)


# ---------------- AUTH ----------------
@router.post("/auth/register")
async def register(body: RegisterBody, response: Response):
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Bu e-posta zaten kayıtlı")
    org = await db.organizations.find_one({}, {"_id": 0})
    org_id = org["id"] if org else new_id("org_")
    if not org:
        await db.organizations.insert_one({"id": org_id, "name": "Fikirizm", "created_at": now_iso()})
    user_id = new_id("user_")
    user = {
        "user_id": user_id,
        "email": email,
        "name": body.name,
        "password_hash": hash_password(body.password),
        "picture": "",
        "org_id": org_id,
        "role": "member",
        "created_at": now_iso(),
    }
    await db.users.insert_one(user)
    ws = await db.workspaces.find_one({"org_id": org_id}, {"_id": 0})
    if ws:
        await db.memberships.update_one(
            {"workspace_id": ws["id"], "user_id": user_id},
            {"$set": {"workspace_id": ws["id"], "user_id": user_id, "org_id": org_id}}, upsert=True)
    access = create_access_token(user_id, email)
    refresh = create_refresh_token(user_id)
    set_auth_cookies(response, access, refresh)
    return {"user": sanitize_user(user), "token": access}


@router.post("/auth/login")
async def login(body: LoginBody, response: Response, request: Request):
    email = body.email.lower()
    ip = request.client.host if request.client else "unknown"
    identifier = f"{ip}:{email}"
    attempt = await db.login_attempts.find_one({"identifier": identifier})
    if attempt and attempt.get("count", 0) >= 5:
        locked_until = attempt.get("locked_until")
        if locked_until and datetime.fromisoformat(locked_until) > datetime.now(timezone.utc):
            raise HTTPException(status_code=429, detail="Çok fazla deneme. 15 dakika sonra tekrar deneyin.")
    user = await db.users.find_one({"email": email})
    if not user or not user.get("password_hash") or not verify_password(body.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": identifier}, {"$inc": {"count": 1}}, upsert=True)
        fresh = await db.login_attempts.find_one({"identifier": identifier})
        if fresh and fresh.get("count", 0) >= 5:
            await db.login_attempts.update_one(
                {"identifier": identifier},
                {"$set": {"locked_until": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()}})
        raise HTTPException(status_code=401, detail="E-posta veya parola hatalı")
    await db.login_attempts.delete_one({"identifier": identifier})
    access = create_access_token(user["user_id"], email)
    refresh = create_refresh_token(user["user_id"])
    set_auth_cookies(response, access, refresh)
    return {"user": sanitize_user(user), "token": access}


@router.post("/auth/session")
async def google_session(body: SessionBody, response: Response):
    async with httpx.AsyncClient() as client_http:
        r = await client_http.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={"X-Session-ID": body.session_id})
    if r.status_code != 200:
        raise HTTPException(status_code=401, detail="Google oturumu doğrulanamadı")
    data = r.json()
    email = data["email"].lower()
    org = await db.organizations.find_one({}, {"_id": 0})
    org_id = org["id"] if org else new_id("org_")
    if not org:
        await db.organizations.insert_one({"id": org_id, "name": "Fikirizm", "created_at": now_iso()})
    user = await db.users.find_one({"email": email})
    if not user:
        user_id = new_id("user_")
        user = {
            "user_id": user_id, "email": email, "name": data.get("name", email),
            "picture": data.get("picture", ""), "org_id": org_id, "role": "member",
            "created_at": now_iso(),
        }
        await db.users.insert_one(user)
        ws = await db.workspaces.find_one({"org_id": org_id}, {"_id": 0})
        if ws:
            await db.memberships.update_one(
                {"workspace_id": ws["id"], "user_id": user_id},
                {"$set": {"workspace_id": ws["id"], "user_id": user_id, "org_id": org_id}}, upsert=True)
    else:
        if data.get("picture"):
            await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"picture": data["picture"]}})
    session_token = data["session_token"]
    await db.user_sessions.insert_one({
        "user_id": user["user_id"], "session_token": session_token,
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "created_at": now_iso(),
    })
    response.set_cookie("session_token", session_token, httponly=True, secure=True,
                        samesite="none", max_age=7 * 24 * 3600, path="/")
    return {"user": sanitize_user(user), "token": session_token}


@router.post("/auth/logout")
async def logout(response: Response, request: Request):
    st = request.cookies.get("session_token")
    if st:
        await db.user_sessions.delete_one({"session_token": st})
    clear_auth_cookies(response)
    return {"ok": True}


@router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


# ---------------- BOOTSTRAP ----------------
@router.get("/bootstrap")
async def bootstrap(user: dict = Depends(get_current_user)):
    org_id = user["org_id"]
    org = await db.organizations.find_one({"id": org_id}, {"_id": 0})
    workspaces = await db.workspaces.find({"org_id": org_id}, {"_id": 0}).to_list(100)
    all_projects = await db.projects.find({"org_id": org_id}, {"_id": 0}).to_list(500)
    if is_privileged(user):
        projects = all_projects
    else:
        projects = [p for p in all_projects if user["user_id"] in p.get("members", []) or p.get("created_by") == user["user_id"]]
    members = await db.users.find({"org_id": org_id}, {"_id": 0, "password_hash": 0}).to_list(200)
    return {"org": org, "workspaces": workspaces, "projects": projects, "members": members,
            "user": user, "templates": {k: {"label": v["label"], "icon": v["icon"]} for k, v in TEMPLATES.items()}}


# ---------------- WORKSPACES ----------------
@router.post("/workspaces")
async def create_workspace(body: WorkspaceBody, user: dict = Depends(get_current_user)):
    ws = {"id": new_id("ws_"), "org_id": user["org_id"], "name": body.name,
          "description": body.description, "created_at": now_iso(), "created_by": user["user_id"]}
    await db.workspaces.insert_one(ws)
    return {k: v for k, v in ws.items() if k != "_id"}


# ---------------- PROJECTS ----------------
@router.post("/projects")
async def create_project(body: ProjectBody, user: dict = Depends(get_current_user)):
    tmpl = get_template(body.template)
    members = list(dict.fromkeys((body.members or []) + [user["user_id"]]))
    proj = {
        "id": new_id("prj_"), "org_id": user["org_id"], "workspace_id": body.workspace_id,
        "name": body.name, "description": body.description, "color": body.color,
        "icon": body.icon or tmpl["icon"], "template": body.template or "general",
        "statuses": tmpl["statuses"], "budget_categories": tmpl["budget_categories"],
        "currency": body.currency or "TRY", "budget_policy": body.budget_policy or "admins",
        "members": members, "created_at": now_iso(), "created_by": user["user_id"],
    }
    await db.projects.insert_one(proj)
    out = {k: v for k, v in proj.items() if k != "_id"}
    await _email_new_members([m for m in members if m != user["user_id"]], user, body.name)
    await log_activity(user["org_id"], body.workspace_id, user, "proje oluşturdu", body.name)
    await manager.broadcast(body.workspace_id, {"type": "project", "action": "create", "data": out})
    return out


@router.patch("/projects/{project_id}")
async def update_project(project_id: str, body: ProjectUpdate, user: dict = Depends(get_current_user)):
    proj = await db.projects.find_one({"id": project_id, "org_id": user["org_id"]}, {"_id": 0})
    if not proj or not can_access_project(user, proj):
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    priv_only = {"members", "budget_policy"}
    if any(k in updates for k in priv_only) and not is_privileged(user):
        for k in priv_only:
            updates.pop(k, None)
    if "members" in updates:
        updates["members"] = list(dict.fromkeys(updates["members"] + [proj.get("created_by")]))
        added = [m for m in updates["members"] if m not in proj.get("members", [])]
    else:
        added = []
    if updates:
        await db.projects.update_one({"id": project_id}, {"$set": updates})
    if added:
        await _email_new_members(added, user, proj["name"])
    proj = await db.projects.find_one({"id": project_id}, {"_id": 0})
    await manager.broadcast(proj["workspace_id"], {"type": "project", "action": "update", "data": proj})
    return proj


@router.delete("/projects/{project_id}")
async def delete_project(project_id: str, user: dict = Depends(get_current_user)):
    proj = await db.projects.find_one({"id": project_id}, {"_id": 0})
    await db.projects.delete_one({"id": project_id, "org_id": user["org_id"]})
    await db.tasks.delete_many({"project_id": project_id})
    if proj:
        await manager.broadcast(proj["workspace_id"], {"type": "project", "action": "delete", "data": {"id": project_id}})
    return {"ok": True}


# ---------------- TASKS ----------------
@router.get("/tasks")
async def list_tasks(user: dict = Depends(get_current_user), project_id: Optional[str] = None,
                     workspace_id: Optional[str] = None, assignee: Optional[str] = None,
                     status: Optional[str] = None, priority: Optional[str] = None,
                     tag: Optional[str] = None, q: Optional[str] = None):
    query = {"org_id": user["org_id"]}
    if project_id:
        proj = await db.projects.find_one({"id": project_id}, {"_id": 0})
        if not can_access_project(user, proj):
            raise HTTPException(status_code=403, detail="Bu projeye erişiminiz yok")
        query["project_id"] = project_id
    else:
        allowed = await accessible_project_ids(user)
        if allowed is not None:
            query["project_id"] = {"$in": allowed}
    if workspace_id:
        query["workspace_id"] = workspace_id
    if assignee:
        query["assignees"] = assignee
    if status:
        query["status"] = status
    if priority:
        query["priority"] = priority
    if tag:
        query["tags"] = tag
    if q:
        query["title"] = {"$regex": re.escape(q), "$options": "i"}
    tasks = await db.tasks.find(query, {"_id": 0}).to_list(1000)
    if not is_privileged(user):
        tasks = [t for t in tasks if can_see_task(user, t)]
    tasks.sort(key=lambda t: t.get("order", 0))
    return tasks


@router.post("/tasks")
async def create_task(body: TaskBody, user: dict = Depends(get_current_user)):
    proj = await db.projects.find_one({"id": body.project_id}, {"_id": 0})
    status = body.status or (proj["statuses"][0]["id"] if proj and proj.get("statuses") else "todo")
    count = await db.tasks.count_documents({"project_id": body.project_id, "status": status})
    task = {
        "id": new_id("tsk_"), "org_id": user["org_id"], "workspace_id": body.workspace_id,
        "project_id": body.project_id, "parent_id": body.parent_id,
        "title": body.title, "description": body.description or "", "status": status,
        "priority": body.priority or "medium", "assignees": body.assignees or [],
        "due_date": body.due_date, "start_date": body.start_date, "tags": body.tags or [],
        "checklist": body.checklist or [], "order": count,
        "visibility": body.visibility or "project", "visible_to": body.visible_to or [],
        "created_at": now_iso(), "created_by": user["user_id"],
    }
    await db.tasks.insert_one(task)
    out = {k: v for k, v in task.items() if k != "_id"}
    for aid in task["assignees"]:
        if aid != user["user_id"]:
            await create_notification(user["org_id"], aid, "assign", f"{user['name']} sizi bir göreve atadı: {task['title']}", f"/proje/{body.project_id}")
    proj_name = proj["name"] if proj else ""
    await _email_new_assignees(task["assignees"], user, task["title"], proj_name)
    await log_activity(user["org_id"], body.workspace_id, user, "görev oluşturdu", body.title)
    await manager.broadcast(body.workspace_id, {"type": "task", "action": "create", "data": out})
    return out


@router.get("/tasks/{task_id}")
async def get_task(task_id: str, user: dict = Depends(get_current_user)):
    task = await db.tasks.find_one({"id": task_id, "org_id": user["org_id"]}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Görev bulunamadı")
    if not can_see_task(user, task):
        raise HTTPException(status_code=403, detail="Bu göreve erişiminiz yok")
    subtasks = await db.tasks.find({"parent_id": task_id}, {"_id": 0}).to_list(500)
    comments = await db.comments.find({"task_id": task_id}, {"_id": 0}).to_list(500)
    comments.sort(key=lambda c: c.get("created_at", ""))
    bitems = await db.budget_items.find({"task_id": task_id}, {"_id": 0}).to_list(300)
    proj = await db.projects.find_one({"id": task["project_id"]}, {"_id": 0})
    bp = sum(float(b.get("planned_amount") or 0) for b in bitems)
    ba = sum(float(b.get("actual_amount") or 0) for b in bitems)
    task["budget_summary"] = {
        "count": len(bitems), "planned": bp, "actual": ba,
        "currency": proj.get("currency", "TRY") if proj else "TRY",
    }
    task["subtasks"] = subtasks
    task["comments"] = comments
    return task


@router.patch("/tasks/{task_id}")
async def update_task(task_id: str, body: TaskUpdate, user: dict = Depends(get_current_user)):
    existing = await db.tasks.find_one({"id": task_id, "org_id": user["org_id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Görev bulunamadı")
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if updates:
        await db.tasks.update_one({"id": task_id}, {"$set": updates})
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if "assignees" in updates:
        for aid in updates["assignees"]:
            if aid not in existing.get("assignees", []) and aid != user["user_id"]:
                await create_notification(user["org_id"], aid, "assign", f"{user['name']} sizi bir göreve atadı: {task['title']}", f"/proje/{task['project_id']}")
        new_assignees = [a for a in updates["assignees"] if a not in existing.get("assignees", [])]
        if new_assignees:
            await _email_new_assignees(new_assignees, user, task["title"])
    if "status" in updates and updates["status"] != existing.get("status"):
        await log_activity(user["org_id"], task["workspace_id"], user, "görev durumunu değiştirdi", task["title"])
    await manager.broadcast(task["workspace_id"], {"type": "task", "action": "update", "data": task})
    return task


@router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, user: dict = Depends(get_current_user)):
    task = await db.tasks.find_one({"id": task_id, "org_id": user["org_id"]}, {"_id": 0})
    await db.tasks.delete_one({"id": task_id})
    await db.tasks.delete_many({"parent_id": task_id})
    if task:
        await manager.broadcast(task["workspace_id"], {"type": "task", "action": "delete", "data": {"id": task_id}})
    return {"ok": True}


@router.post("/tasks/bulk")
async def bulk_update(body: BulkUpdate, user: dict = Depends(get_current_user)):
    allowed = {"status", "priority", "tags", "assignees", "due_date"}
    updates = {k: v for k, v in body.updates.items() if k in allowed}
    if updates:
        await db.tasks.update_many({"id": {"$in": body.ids}, "org_id": user["org_id"]}, {"$set": updates})
    tasks = await db.tasks.find({"id": {"$in": body.ids}}, {"_id": 0}).to_list(1000)
    if tasks:
        await manager.broadcast(tasks[0]["workspace_id"], {"type": "task", "action": "bulk", "data": tasks})
    return tasks


@router.post("/tasks/{task_id}/comments")
async def add_task_comment(task_id: str, body: CommentBody, user: dict = Depends(get_current_user)):
    task = await db.tasks.find_one({"id": task_id, "org_id": user["org_id"]}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Görev bulunamadı")
    comment = {
        "id": new_id("cmt_"), "task_id": task_id, "user_id": user["user_id"],
        "user_name": user["name"], "user_picture": user.get("picture", ""),
        "text": body.text, "created_at": now_iso(),
    }
    await db.comments.insert_one(comment)
    out = {k: v for k, v in comment.items() if k != "_id"}
    for aid in task.get("assignees", []):
        if aid != user["user_id"]:
            await create_notification(user["org_id"], aid, "comment", f"{user['name']} bir göreve yorum yaptı: {task['title']}", f"/proje/{task['project_id']}")
    await manager.broadcast(task["workspace_id"], {"type": "comment", "action": "create", "data": out})
    return out


# ---------------- IDEAS ----------------
@router.get("/ideas")
async def list_ideas(user: dict = Depends(get_current_user), workspace_id: Optional[str] = None,
                     sort: str = "votes"):
    query = {"org_id": user["org_id"]}
    if workspace_id:
        query["workspace_id"] = workspace_id
    ideas = await db.ideas.find(query, {"_id": 0}).to_list(1000)
    for i in ideas:
        i["vote_count"] = len(i.get("upvotes", []))
        i["comment_count"] = await db.comments.count_documents({"idea_id": i["id"]})
    if sort == "newest":
        ideas.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    elif sort == "discussed":
        ideas.sort(key=lambda x: x.get("comment_count", 0), reverse=True)
    else:
        ideas.sort(key=lambda x: x.get("vote_count", 0), reverse=True)
    return ideas


@router.post("/ideas")
async def create_idea(body: IdeaBody, user: dict = Depends(get_current_user)):
    idea = {
        "id": new_id("idea_"), "org_id": user["org_id"], "workspace_id": body.workspace_id,
        "project_id": body.project_id, "title": body.title, "description": body.description or "",
        "status": "new", "upvotes": [], "converted_task_id": None,
        "created_at": now_iso(), "created_by": user["user_id"], "created_by_name": user["name"],
    }
    await db.ideas.insert_one(idea)
    out = {k: v for k, v in idea.items() if k != "_id"}
    out["vote_count"] = 0
    out["comment_count"] = 0
    await log_activity(user["org_id"], body.workspace_id, user, "fikir ekledi", body.title)
    await manager.broadcast(body.workspace_id, {"type": "idea", "action": "create", "data": out})
    return out


@router.patch("/ideas/{idea_id}")
async def update_idea(idea_id: str, body: IdeaUpdate, user: dict = Depends(get_current_user)):
    existing = await db.ideas.find_one({"id": idea_id, "org_id": user["org_id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Fikir bulunamadı")
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if updates:
        await db.ideas.update_one({"id": idea_id}, {"$set": updates})
    idea = await db.ideas.find_one({"id": idea_id}, {"_id": 0})
    idea["vote_count"] = len(idea.get("upvotes", []))
    if "status" in updates and updates["status"] != existing.get("status"):
        if existing.get("created_by") != user["user_id"]:
            await create_notification(user["org_id"], existing["created_by"], "idea", f"Fikrinizin durumu güncellendi: {idea['title']}", "/fikirler")
        await log_activity(user["org_id"], idea["workspace_id"], user, "fikir durumunu değiştirdi", idea["title"])
    await manager.broadcast(idea["workspace_id"], {"type": "idea", "action": "update", "data": idea})
    return idea


@router.post("/ideas/{idea_id}/vote")
async def vote_idea(idea_id: str, user: dict = Depends(get_current_user)):
    idea = await db.ideas.find_one({"id": idea_id, "org_id": user["org_id"]}, {"_id": 0})
    if not idea:
        raise HTTPException(status_code=404, detail="Fikir bulunamadı")
    upvotes = idea.get("upvotes", [])
    if user["user_id"] in upvotes:
        upvotes.remove(user["user_id"])
    else:
        upvotes.append(user["user_id"])
        if idea.get("created_by") != user["user_id"]:
            await create_notification(user["org_id"], idea["created_by"], "vote", f"{user['name']} fikrinizi oyladı: {idea['title']}", "/fikirler")
    await db.ideas.update_one({"id": idea_id}, {"$set": {"upvotes": upvotes}})
    idea["upvotes"] = upvotes
    idea["vote_count"] = len(upvotes)
    await manager.broadcast(idea["workspace_id"], {"type": "idea", "action": "update", "data": idea})
    return idea


@router.get("/ideas/{idea_id}/comments")
async def get_idea_comments(idea_id: str, user: dict = Depends(get_current_user)):
    comments = await db.comments.find({"idea_id": idea_id}, {"_id": 0}).to_list(500)
    comments.sort(key=lambda c: c.get("created_at", ""))
    return comments


@router.post("/ideas/{idea_id}/comments")
async def add_idea_comment(idea_id: str, body: CommentBody, user: dict = Depends(get_current_user)):
    idea = await db.ideas.find_one({"id": idea_id, "org_id": user["org_id"]}, {"_id": 0})
    if not idea:
        raise HTTPException(status_code=404, detail="Fikir bulunamadı")
    comment = {
        "id": new_id("cmt_"), "idea_id": idea_id, "user_id": user["user_id"],
        "user_name": user["name"], "user_picture": user.get("picture", ""),
        "text": body.text, "created_at": now_iso(),
    }
    await db.comments.insert_one(comment)
    out = {k: v for k, v in comment.items() if k != "_id"}
    await manager.broadcast(idea["workspace_id"], {"type": "idea_comment", "action": "create", "data": out})
    return out


@router.post("/ideas/{idea_id}/convert")
async def convert_idea(idea_id: str, user: dict = Depends(get_current_user)):
    idea = await db.ideas.find_one({"id": idea_id, "org_id": user["org_id"]}, {"_id": 0})
    if not idea:
        raise HTTPException(status_code=404, detail="Fikir bulunamadı")
    project_id = idea.get("project_id")
    if not project_id:
        proj = await db.projects.find_one({"workspace_id": idea["workspace_id"]}, {"_id": 0})
        if not proj:
            raise HTTPException(status_code=400, detail="Dönüştürülecek bir proje bulunamadı")
        project_id = proj["id"]
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    first_status = project["statuses"][0]["id"] if project and project.get("statuses") else "todo"
    task = {
        "id": new_id("tsk_"), "org_id": user["org_id"], "workspace_id": idea["workspace_id"],
        "project_id": project_id, "parent_id": None, "title": idea["title"],
        "description": idea.get("description", ""), "status": first_status, "priority": "medium",
        "assignees": [], "due_date": None, "start_date": None, "tags": ["fikirden"],
        "checklist": [], "order": 0, "created_at": now_iso(), "created_by": user["user_id"],
    }
    await db.tasks.insert_one(task)
    await db.ideas.update_one({"id": idea_id}, {"$set": {"status": "approved", "converted_task_id": task["id"]}})
    out = {k: v for k, v in task.items() if k != "_id"}
    await log_activity(user["org_id"], idea["workspace_id"], user, "fikri göreve dönüştürdü", idea["title"])
    await manager.broadcast(idea["workspace_id"], {"type": "task", "action": "create", "data": out})
    return {"task": out, "project_id": project_id}


# ---------------- DASHBOARD ----------------
@router.get("/dashboard")
async def dashboard(user: dict = Depends(get_current_user), workspace_id: Optional[str] = None):
    query = {"org_id": user["org_id"]}
    if workspace_id:
        query["workspace_id"] = workspace_id
    allowed = await accessible_project_ids(user)
    if allowed is not None:
        query["project_id"] = {"$in": allowed}
    projects = await db.projects.find({"org_id": user["org_id"]}, {"_id": 0}).to_list(2000)
    pmap = {p["id"]: p for p in projects}

    def status_meta(t):
        proj = pmap.get(t.get("project_id"))
        statuses = proj.get("statuses", []) if proj else []
        for s in statuses:
            if s["id"] == t.get("status"):
                return s
        return {"id": t.get("status", "todo"), "name": t.get("status", "todo"), "color": "#71717A", "done": False}

    tasks = await db.tasks.find({**query, "parent_id": None}, {"_id": 0}).to_list(2000)
    now = datetime.now(timezone.utc)
    week_end = now + timedelta(days=7)

    def is_done(t):
        return bool(status_meta(t).get("done"))

    open_tasks = [t for t in tasks if not is_done(t)]
    overdue, upcoming = [], []
    for t in open_tasks:
        if t.get("due_date"):
            try:
                dd = datetime.fromisoformat(t["due_date"].replace("Z", "+00:00"))
                if dd.tzinfo is None:
                    dd = dd.replace(tzinfo=timezone.utc)
                if dd < now:
                    overdue.append(t)
                elif dd <= week_end:
                    upcoming.append(t)
            except Exception:
                pass
    dist = {}
    for t in tasks:
        sm = status_meta(t)
        key = sm["name"]
        if key not in dist:
            dist[key] = {"name": key, "value": 0, "color": sm.get("color", "#71717A")}
        dist[key]["value"] += 1
    workload = {}
    for t in open_tasks:
        for a in t.get("assignees", []):
            workload[a] = workload.get(a, 0) + 1
    my_tasks = [t for t in open_tasks if user["user_id"] in t.get("assignees", [])]
    activities = await db.activities.find(query if allowed is None or "project_id" not in query else {"org_id": user["org_id"], **({"workspace_id": workspace_id} if workspace_id else {})}, {"_id": 0}).to_list(500)
    activities.sort(key=lambda a: a.get("created_at", ""), reverse=True)
    return {
        "open_count": len(open_tasks),
        "overdue_count": len(overdue),
        "upcoming_count": len(upcoming),
        "total_count": len(tasks),
        "done_count": len([t for t in tasks if is_done(t)]),
        "status_distribution": list(dist.values()),
        "workload": workload,
        "my_tasks": my_tasks[:10],
        "overdue_tasks": overdue[:10],
        "upcoming_tasks": sorted(upcoming, key=lambda t: t.get("due_date", ""))[:10],
        "recent_activities": activities[:15],
    }


# ---------------- NOTIFICATIONS ----------------
@router.get("/notifications")
async def get_notifications(user: dict = Depends(get_current_user)):
    notes = await db.notifications.find({"user_id": user["user_id"]}, {"_id": 0}).to_list(200)
    notes.sort(key=lambda n: n.get("created_at", ""), reverse=True)
    return notes


@router.post("/notifications/{note_id}/read")
async def read_notification(note_id: str, user: dict = Depends(get_current_user)):
    await db.notifications.update_one({"id": note_id, "user_id": user["user_id"]}, {"$set": {"read": True}})
    return {"ok": True}


@router.post("/notifications/read-all")
async def read_all(user: dict = Depends(get_current_user)):
    await db.notifications.update_many({"user_id": user["user_id"]}, {"$set": {"read": True}})
    return {"ok": True}


# ---------------- MEMBERS ----------------
@router.get("/members")
async def get_members(user: dict = Depends(get_current_user)):
    members = await db.users.find({"org_id": user["org_id"]}, {"_id": 0, "password_hash": 0}).to_list(200)
    return members


@router.post("/members/invite")
async def invite_member(body: InviteBody, user: dict = Depends(get_current_user)):
    if user.get("role") not in ("owner", "admin"):
        raise HTTPException(status_code=403, detail="Üye davet etme yetkiniz yok")
    email = body.email.lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Bu e-posta zaten üye")
    import os
    user_id = new_id("user_")
    new_user = {
        "user_id": user_id, "email": email, "name": body.name,
        "password_hash": hash_password(os.environ.get("DEMO_PASSWORD", "Demo2025!")),
        "picture": "", "org_id": user["org_id"], "role": body.role, "created_at": now_iso(),
    }
    await db.users.insert_one(new_user)
    return sanitize_user(new_user)


# ---------------- SEARCH ----------------
@router.get("/search")
async def search(q: str, user: dict = Depends(get_current_user)):
    if not q or len(q) < 1:
        return {"tasks": [], "ideas": []}
    tquery = {"org_id": user["org_id"], "title": {"$regex": re.escape(q), "$options": "i"}}
    allowed = await accessible_project_ids(user)
    if allowed is not None:
        tquery["project_id"] = {"$in": allowed}
    tasks = await db.tasks.find(tquery, {"_id": 0}).to_list(20)
    ideas = await db.ideas.find(
        {"org_id": user["org_id"], "title": {"$regex": re.escape(q), "$options": "i"}}, {"_id": 0}).to_list(20)
    return {"tasks": tasks, "ideas": ideas}


# ---------------- ACTIVITIES ----------------
@router.get("/activities")
async def get_activities(user: dict = Depends(get_current_user), workspace_id: Optional[str] = None):
    query = {"org_id": user["org_id"]}
    if workspace_id:
        query["workspace_id"] = workspace_id
    acts = await db.activities.find(query, {"_id": 0}).to_list(200)
    acts.sort(key=lambda a: a.get("created_at", ""), reverse=True)
    return acts[:50]


# ---------------- BUDGET ----------------
def _budget_summary(items):
    s = {"planned_income": 0.0, "actual_income": 0.0, "planned_expense": 0.0, "actual_expense": 0.0}
    cats = {}
    for it in items:
        p = float(it.get("planned_amount") or 0)
        a = float(it.get("actual_amount") or 0)
        if it["type"] == "income":
            s["planned_income"] += p
            s["actual_income"] += a
        else:
            s["planned_expense"] += p
            s["actual_expense"] += a
        key = (it["type"], it.get("category", "Diğer"))
        c = cats.setdefault(key, {"type": it["type"], "category": it.get("category", "Diğer"), "planned": 0.0, "actual": 0.0})
        c["planned"] += p
        c["actual"] += a
    s["planned_balance"] = s["planned_income"] - s["planned_expense"]
    s["actual_balance"] = s["actual_income"] - s["actual_expense"]
    s["by_category"] = list(cats.values())
    return s


@router.get("/projects/{project_id}/budget")
async def get_budget(project_id: str, user: dict = Depends(get_current_user)):
    proj = await db.projects.find_one({"id": project_id, "org_id": user["org_id"]}, {"_id": 0})
    if not proj or not can_access_project(user, proj):
        raise HTTPException(status_code=403, detail="Bu projeye erişiminiz yok")
    items = await db.budget_items.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    items.sort(key=lambda x: x.get("date") or x.get("created_at", ""), reverse=True)
    return {
        "items": items,
        "summary": _budget_summary(items),
        "currency": proj.get("currency", "TRY"),
        "categories": proj.get("budget_categories", {"income": [], "expense": []}),
        "can_edit": can_edit_budget(user, proj),
        "policy": proj.get("budget_policy", "admins"),
    }


@router.post("/projects/{project_id}/budget")
async def create_budget_item(project_id: str, body: BudgetBody, user: dict = Depends(get_current_user)):
    proj = await db.projects.find_one({"id": project_id, "org_id": user["org_id"]}, {"_id": 0})
    if not proj or not can_edit_budget(user, proj):
        raise HTTPException(status_code=403, detail="Bütçeyi düzenleme yetkiniz yok")
    item = {
        "id": new_id("bdg_"), "org_id": user["org_id"], "workspace_id": proj["workspace_id"],
        "project_id": project_id, "type": body.type, "category": body.category,
        "description": body.description or "", "planned_amount": body.planned_amount or 0,
        "actual_amount": body.actual_amount or 0, "date": body.date, "responsible": body.responsible,
        "task_id": body.task_id, "created_at": now_iso(), "created_by": user["user_id"],
    }
    await db.budget_items.insert_one(item)
    out = {k: v for k, v in item.items() if k != "_id"}
    await log_activity(user["org_id"], proj["workspace_id"], user, "bütçe kalemi ekledi", body.category)
    await manager.broadcast(proj["workspace_id"], {"type": "budget", "action": "create", "data": out})
    return out


@router.patch("/budget/{item_id}")
async def update_budget_item(item_id: str, body: BudgetUpdate, user: dict = Depends(get_current_user)):
    item = await db.budget_items.find_one({"id": item_id, "org_id": user["org_id"]}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Bütçe kalemi bulunamadı")
    proj = await db.projects.find_one({"id": item["project_id"]}, {"_id": 0})
    if not can_edit_budget(user, proj):
        raise HTTPException(status_code=403, detail="Bütçeyi düzenleme yetkiniz yok")
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if updates:
        await db.budget_items.update_one({"id": item_id}, {"$set": updates})
    item = await db.budget_items.find_one({"id": item_id}, {"_id": 0})
    await manager.broadcast(proj["workspace_id"], {"type": "budget", "action": "update", "data": item})
    return item


@router.delete("/budget/{item_id}")
async def delete_budget_item(item_id: str, user: dict = Depends(get_current_user)):
    item = await db.budget_items.find_one({"id": item_id, "org_id": user["org_id"]}, {"_id": 0})
    if not item:
        return {"ok": True}
    proj = await db.projects.find_one({"id": item["project_id"]}, {"_id": 0})
    if not can_edit_budget(user, proj):
        raise HTTPException(status_code=403, detail="Bütçeyi düzenleme yetkiniz yok")
    await db.budget_items.delete_one({"id": item_id})
    await manager.broadcast(proj["workspace_id"], {"type": "budget", "action": "delete", "data": {"id": item_id}})
    return {"ok": True}


@router.get("/projects/{project_id}/budget/export")
async def export_budget(project_id: str, fmt: str = "xlsx", user: dict = Depends(get_current_user)):
    proj = await db.projects.find_one({"id": project_id, "org_id": user["org_id"]}, {"_id": 0})
    if not proj or not can_access_project(user, proj):
        raise HTTPException(status_code=403, detail="Bu projeye erişiminiz yok")
    items = await db.budget_items.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    members = await db.users.find({"org_id": user["org_id"]}, {"_id": 0}).to_list(200)
    tasks = await db.tasks.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    mmap = {m["user_id"]: m.get("name", "") for m in members}
    tmap = {t["id"]: t.get("title", "") for t in tasks}
    cur = proj.get("currency", "TRY")
    s = _budget_summary(items)
    rows = [("Tür", "Kategori", "Açıklama", "Planlanan", "Gerçekleşen", "Tarih", "Sorumlu", "Görev")]
    for it in items:
        rows.append((
            "Gelir" if it["type"] == "income" else "Gider", it.get("category", ""),
            it.get("description", ""), float(it.get("planned_amount") or 0),
            float(it.get("actual_amount") or 0), (it.get("date") or "")[:10],
            mmap.get(it.get("responsible"), ""), tmap.get(it.get("task_id"), ""),
        ))
    fname = f"butce_{proj['name']}".replace(" ", "_")

    if fmt == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        font = "Helvetica"
        for path in ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                     "/usr/share/fonts/dejavu/DejaVuSans.ttf"):
            if os.path.exists(path):
                try:
                    pdfmetrics.registerFont(TTFont("DejaVu", path))
                    font = "DejaVu"
                    break
                except Exception:
                    pass

        def safe(v):
            t = str(v)
            return t if font == "DejaVu" else t.encode("latin-1", "replace").decode("latin-1")

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=18 * mm)
        styles = getSampleStyleSheet()
        styles["Title"].fontName = font
        styles["Normal"].fontName = font
        elems = [Paragraph(safe(f"{proj['name']} — Bütçe Raporu ({cur})"), styles["Title"]), Spacer(1, 8)]
        summ = (f"Toplam Gelir (Gerç.): {s['actual_income']:.0f} / Plan {s['planned_income']:.0f}   |   "
                f"Toplam Gider (Gerç.): {s['actual_expense']:.0f} / Plan {s['planned_expense']:.0f}   |   "
                f"Bakiye (Gerç.): {s['actual_balance']:.0f}")
        elems += [Paragraph(safe(summ), styles["Normal"]), Spacer(1, 10)]
        tdata = [[safe(c) for c in rows[0]]] + [[safe(c) for c in r] for r in rows[1:]]
        table = Table(tdata, repeatRows=1)
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d4d4d8")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f4f5")]),
        ]))
        elems.append(table)
        doc.build(elems)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": f'attachment; filename="{fname}.pdf"'})

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Bütçe"
    ws.append([f"{proj['name']} — Bütçe ({cur})"])
    ws.append([f"Bakiye (Gerç.): {s['actual_balance']:.0f}  |  Gelir: {s['actual_income']:.0f}  |  Gider: {s['actual_expense']:.0f}"])
    ws.append([])
    ws.append(list(rows[0]))
    hdr_row = ws.max_row
    for r in rows[1:]:
        ws.append(list(r))
    for cell in ws[hdr_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F46E5")
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 45)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'})


# ---------------- WEBSOCKET ----------------
@router.websocket("/ws/{workspace_id}")
async def websocket_endpoint(websocket: WebSocket, workspace_id: str):
    user = await get_ws_user(websocket)
    if not user:
        await websocket.close(code=1008)
        return
    await manager.connect(workspace_id, websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(workspace_id, websocket)
    except Exception:
        manager.disconnect(workspace_id, websocket)
